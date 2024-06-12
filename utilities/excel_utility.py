import openpyxl
from openpyxl.styles import PatternFill


def GetRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def GetColumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)



def GetReaddata(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,columnno).value

def writedata(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)



def fillGreencolor(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenfill
    workbook.save(file)